import time
import os
import openpyxl  
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

def descargar_archivo_bvc():
    ruta_script = os.path.dirname(os.path.abspath(__file__))
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    prefs = {"download.default_directory": ruta_script, "download.prompt_for_download": False}
    chrome_options.add_experimental_option("prefs", prefs)
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    try:
        driver.get("https://www.bvc.com.co/mercado-local-en-linea?tab=renta-fija_deuda-corporativa")
        driver.maximize_window()
        time.sleep(15)
        driver.execute_script("document.querySelectorAll('.ant-popover, .ant-modal, .ant-modal-mask').forEach(el => el.remove());")
        
        script_clic = """
            const btn = Array.from(document.querySelectorAll('button')).find(b => b.innerText.includes('Descarga'));
            if(btn) { btn.scrollIntoView({block: 'center'}); btn.click(); return true; }
            return false;
        """
        if driver.execute_script(script_clic):
            time.sleep(10)
            return True
        return False
    except:
        return False
    finally:
        driver.quit()

def procesar_logica_excel():
    nombre_excel = "Renta fija.xlsm"
    
    if not os.path.exists(nombre_excel):
        print(f"No se encontró el archivo: {nombre_excel}")
        return

    
    wb = openpyxl.load_workbook(nombre_excel, keep_vba=True) 
    if 'Fechas' not in wb.sheetnames:
        print("No se encontró la hoja 'Fechas'")
        return
    
    ws = wb['Fechas']
    hoy_dt = datetime.now()
    fecha_hoy_str = hoy_dt.strftime("%#d/%m/%Y") 
    
    print(f"Buscando hoy: {fecha_hoy_str}")
    
    encontrado = False
    # Iterar sobre las filas asume encabezado
    for row in range(2, ws.max_row + 1):
        celda_fecha = ws.cell(row=row, column=1).value
        
        # Convertir fecha de la celda a string para comparacion
        fecha_celda_str = ""
        if isinstance(celda_fecha, datetime):
            fecha_celda_str = celda_fecha.strftime("%#d/%m/%Y")
        else:
            fecha_celda_str = str(celda_fecha).split()[0].replace("-", "/")

        if fecha_celda_str == fecha_hoy_str:
            encontrado = True
            dia_bursatil = ws.cell(row=row, column=2).value
            estado_actual = ws.cell(row=row, column=3).value
            
            if str(estado_actual).lower() == 'ejecutado':
                print("Ya fue ejecutado hoy.")
                return

            if str(dia_bursatil).upper() == 'SI':
                print("Día bursátil detectado. Descargando...")
                if descargar_archivo_bvc():
                    ws.cell(row=row, column=3).value = "Ejecutado"
            else:
                print("Hoy NO es bursátil según Excel.")
                ws.cell(row=row, column=3).value = "Ejecutado"
            break

    if encontrado:
        wb.save(nombre_excel) # Guarda manteniendo el formato .xlsm 
        print("Archivo .xlsm actualizado y macros preservadas.")
    else:
        print("Fecha de hoy no encontrada en el Excel.")

if __name__ == "__main__":
    procesar_logica_excel()